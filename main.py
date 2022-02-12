from config import token_b, s_k_email, s_k_id, bot_email, bot_password, path_dir, passwordc
from telebot import types
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from datetime import date, timedelta
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.text import MIMEText

import os
import shutil
import smtplib
import os.path
import datetime
import telebot
import traceback
import time

# main data
os.chdir(path_dir)
bot = telebot.TeleBot(token_b, threaded=False)
user_dict = {}
path_to_n_cont = 'n_count.txt'
skip_list = []

now = datetime.datetime.now()
date_of_day = date.today()
path_of_registration_file = 'registration_file.txt'
path_of_registration_file_admins = 'registration_file_admins.txt'
password_of_admins = passwordc # main variable

# Keyboard to confirm the payment
inline_keyboard_paid = types.InlineKeyboardMarkup()
paid_key_1 = types.InlineKeyboardButton(text='Да, безнал', callback_data='paid_1')
# Cash payment button
paid_key_2 = types.InlineKeyboardButton(text='Да, наличными', callback_data='paid_2')
# button not pay
paid_key_3 = types.InlineKeyboardButton(text='Нет', callback_data='paid_3')
inline_keyboard_paid.add(paid_key_1, paid_key_2, paid_key_3)

# keyboard check order
inline_keyboard_order_check = types.InlineKeyboardMarkup()
order_check_1 = types.InlineKeyboardButton(text='Да', callback_data='order_check_yes')
order_check_2 = types.InlineKeyboardButton(text='Нет', callback_data='order_check_no')
inline_keyboard_order_check.add(order_check_1, order_check_2)

# keyboard check spare
spare_confirm = types.InlineKeyboardMarkup()
spare_check_1 = types.InlineKeyboardButton(text='Да', callback_data='spare_check_yes')
spare_check_2 = types.InlineKeyboardButton(text='Нет', callback_data='spare_check_no')
spare_confirm.add(spare_check_1, spare_check_2)


# kb for admins
def kb_cmd_admin():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    registration = types.KeyboardButton('Регистрация')
    add_org = types.KeyboardButton('Добавить организацию')
    del_org = types.KeyboardButton('Удалить организацию')
    add_worker = types.KeyboardButton('Добавить работника')
    del_worker = types.KeyboardButton('Удалить работника')
    refresh_storage_ = types.KeyboardButton('Обновить склад')
    del_spare = types.KeyboardButton('Удалить запчасть')
    code = types.KeyboardButton('Ввести коды')
    back = types.KeyboardButton('Вернуться')

    markup.add(del_spare, refresh_storage_, code)
    markup.add(add_org, del_org)
    markup.add(add_worker, del_worker)
    markup.add(registration, back)
    return markup

def code_giver_kb(row):
    kb = types.InlineKeyboardMarkup(row_width=2)
    skip = types.InlineKeyboardButton(text='Пропустить', callback_data=f'skip_row_{row}')
    kb.add(skip, cancel)
    return kb

def get_job_category():
    kb = types.InlineKeyboardMarkup(row_width=4)
    locksmith_job = types.InlineKeyboardButton(text='Слесарные работы', callback_data=f'cate_locksmith')
    pickup_job = types.InlineKeyboardButton(text='Кузовные работы', callback_data=f'cate_pickupjob')
    sticker = types.InlineKeyboardButton(text='Оклейка авто', callback_data=f'cate_sticker')
    wash = types.InlineKeyboardButton(text='Мойка авто', callback_data=f'cate_wash')
    kb.add(locksmith_job, pickup_job)
    kb.add(sticker, wash)
    return kb

def bu():
    kb = types.InlineKeyboardMarkup(row_width=2)
    used = types.InlineKeyboardButton(text='Б/у', callback_data='used_spare')
    new = types.InlineKeyboardButton(text='Н', callback_data='new_spare')
    kb.add(used, new)
    return kb

def car_brand():
    kb = types.InlineKeyboardMarkup(row_width=3)
    skoda = types.InlineKeyboardButton(text='Ш', callback_data='brand_skoda')
    kia = types.InlineKeyboardButton(text='К', callback_data='brand_kia')
    both = types.InlineKeyboardButton(text='Ш,К', callback_data='brand_skoda_kia')
    kb.add(skoda, kia, both)
    return kb
# kb with admin's name
def admins_keyboard():
    inline_keyboard_admins = types.InlineKeyboardMarkup(row_width=2)
    with open(path_of_registration_file_admins, 'r') as f:
        lines = f.readlines()
        for line in lines:
            if line != None and line != '\n':
                list_line = list(line.split())
                admin = types.InlineKeyboardButton(text=f'{list_line[1]}', callback_data=f'admin_{list_line[0]}')
                inline_keyboard_admins.add(admin)
    return inline_keyboard_admins


# org kb
def org_keyboard(to_use):
    inline_keyboard_org = types.InlineKeyboardMarkup(row_width=4)
    try:
        with open('organiztion/list_of_org.txt', 'r') as f:
            row = []
            for line in f:
                if line != None and line != '\n':
                    if to_use:
                        call_back = 'org_' + line
                    else:
                        call_back = 'del_org_' + line
                    row.append(types.InlineKeyboardButton(text=line, callback_data=call_back))
                    if len(row) >= 4:
                        inline_keyboard_org.add(*row)
                        row = []
            if len(row) > 0:
                inline_keyboard_org.add(*row)
            inline_keyboard_org.add(cancel)
        return inline_keyboard_org
    except:
        pass


def spare_keyboard_get(to_use):
    il_kb_spare = types.InlineKeyboardMarkup(row_width=4)
    # del_spare_ give_ get_

    xlsx_btn_name = types.InlineKeyboardButton(text='По ключевому слову', callback_data=f'xlsx_{to_use}_name')
    xlsx_btn_code = types.InlineKeyboardButton(text='По коду', callback_data=f'xlsx_{to_use}_code')
    il_kb_spare.add(xlsx_btn_name)
    il_kb_spare.add(xlsx_btn_code)
    il_kb_spare.add(cancel)
    return il_kb_spare


############################################################################################################################
# worker_keyboard
def worker_keyboard(to_use_worker):
    inline_keyboard_worker = types.InlineKeyboardMarkup(row_width=3)
    try:
        with open('workers.txt', 'r') as f:
            row = []
            for line in f:
                if line != None and line != '\n':
                    if to_use_worker:
                        call_back = 'worker_' + line
                    else:
                        call_back = 'del_worker_' + line
                    row.append(types.InlineKeyboardButton(text=line, callback_data=call_back))
                    if len(row) >= 3:
                        inline_keyboard_worker.add(*row)
                        row = []
            if len(row) > 0:
                inline_keyboard_worker.add(*row)
            inline_keyboard_worker.add(cancel)
        return inline_keyboard_worker
    except:
        pass


################################################################################################################################
# kb confirm order
inline_keyboard_send_order = types.InlineKeyboardMarkup()
send_check_1 = types.InlineKeyboardButton(text='Да', callback_data='send_check_yes')
send_check_2 = types.InlineKeyboardButton(text='Нет', callback_data='send_check_no')
inline_keyboard_send_order.add(send_check_1, send_check_2)

# reg kb
inline_keyboard_reg = types.InlineKeyboardMarkup()
reg_1 = types.InlineKeyboardButton(text='Да', callback_data='reg_yes')
reg_2 = types.InlineKeyboardButton(text='Нет', callback_data='reg_no')
inline_keyboard_reg.add(reg_1, reg_2)

# kb change reg data
inline_keyboard_reg_change = types.InlineKeyboardMarkup()
change_reg_1 = types.InlineKeyboardButton(text='Изменить', callback_data='change_reg_yes')
change_reg_2 = types.InlineKeyboardButton(text='Удалить', callback_data='change_reg_no')
cancel = types.InlineKeyboardButton(text='Отмена', callback_data='cancel')
inline_keyboard_reg_change.add(change_reg_1, change_reg_2, )
inline_keyboard_reg_change.add(cancel)

# kb choose change data
inline_keyboard_change_name = types.InlineKeyboardMarkup()
change_n_1 = types.InlineKeyboardButton(text='Имя', callback_data='cn_name')
change_n_2 = types.InlineKeyboardButton(text='Фамилию', callback_data='cn_second_name')
change_n_3 = types.InlineKeyboardButton(text='Отчество', callback_data='cn_thierd_name')
inline_keyboard_change_name.add(change_n_1, change_n_2, change_n_3)
inline_keyboard_change_name.add(cancel)

# kb choose change data
# kb org yes no continue
inline_keyboard_org_continue = types.InlineKeyboardMarkup()
org_con_yes = types.InlineKeyboardButton(text='Да', callback_data='org_yes')
org_con_no = types.InlineKeyboardButton(text='Нет', callback_data='org_no')
inline_keyboard_org_continue.add(org_con_yes, org_con_no)


# get fio from files
def fio_getter(call):
    if check_id_in_file(call.message.chat.id, path_of_registration_file_admins):
        path = path_of_registration_file_admins
    else:
        path = path_of_registration_file
    with open(path, 'r') as f:
        all_line = f.readlines()
        for line in all_line:
            if str(call.message.chat.id) in line:
                list_of_line = list(line.split())
                fio_user = f'{list_of_line[2].capitalize()} {list_of_line[1][0].upper()}.{list_of_line[3][0].upper()}.'
                return fio_user


# check whether the auto is in organization file excel
def excel_check(name, call):
    user_dict[call.message.chat.id]['organization'] = name[12:]
    if os.path.exists(f'{name}.xlsx'):
        try:
            wb = load_workbook(f'{name}.xlsx')
            ws = wb.active
            number_column = ws['C']
            for cell in number_column:
                if user_dict[call.message.chat.id]['number_auto'].lower() in cell.value.lower():
                    wb.save(f'{name}.xlsx')
                    return True
            wb.save(f'{name}.xlsx')
            return False
        except:
            pass


# edit excel file
def excel_maker(call):
    with open(path_to_n_cont, 'r') as f:
        n = int(f.read().strip())
    shapka = ['Номер', 'Время', 'ФИО сотрудника', 'Наличные/безналичные', 'Номер авто', 'Организация', 'Цена',
              'Вид работ', 'Пробег']
    list_of_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    n_local = int(user_dict[call.message.chat.id]['local_n']) + 1
    if os.path.exists('excel_files/report.xlsx'):
        wb_obj = load_workbook('excel_files/report.xlsx')
        sheet_obj = wb_obj.active
        sheet_obj[f'A{n_local}'] = int(n_local) - 1
        sheet_obj[f'B{n_local}'] = str(datetime.datetime.today().strftime("%H:%M:%S"))
        # sheet_obj[f'C{n_local}'] = fio_getter(call)
        sheet_obj[f'C{n_local}'] = user_dict[call.message.chat.id]['name']
        sheet_obj[f'D{n_local}'] = user_dict[call.message.chat.id]['paid']
        sheet_obj_D = sheet_obj[f'D{n_local}']
        sheet_obj[f'E{n_local}'] = user_dict[call.message.chat.id]['number_auto'].upper()
        sheet_obj[f'F{n_local}'] = user_dict[call.message.chat.id]['organization']
        sheet_obj[f'G{n_local}'] = user_dict[call.message.chat.id]['price']
        sheet_obj[f'H{n_local}'] = user_dict[call.message.chat.id]['job_type']
        sheet_obj[f'I{n_local}'] = user_dict[call.message.chat.id]['probeg']

        wb_obj.save('excel_files/report.xlsx')
        with open(path_to_n_cont, 'w') as f:
            f.write(str(n + 1))
    else:
        # create excel file if it don't exist
        wb = Workbook()
        sheet = wb.active
        for i in range(0, 9):
            sheet[f'{list_of_letters[i]}1'] = shapka[i]
        wb.save('excel_files/report.xlsx')
        excel_maker(call)


# button click handler
@bot.callback_query_handler(func=lambda call: True)
def callback_worker(call):
    global paid, send_photo, job_type
    # paid button
    if call.data[:4] == "paid":
        if call.data[-1] == '1':
            user_dict[call.message.chat.id]['paid'] = "Безнал"
            bot.send_message(call.message.chat.id, "Введите организацию, которая производит оплату",
                             reply_markup=org_keyboard(True))


        elif call.data[-1] == '2':
            user_dict[call.message.chat.id]['paid'] = "Нал"
            bot.send_message(chat_id=call.message.chat.id,
                             text="Выберете текущего администратора",
                             reply_markup=admins_keyboard())
            user_dict[call.message.chat.id]['organization'] = 'Нет'

        else:
            user_dict[call.message.chat.id]['paid'] = "Не оплачено"
            bot.send_message(chat_id=call.message.chat.id,
                             text="Выберете текущего администратора",
                             reply_markup=admins_keyboard())
            user_dict[call.message.chat.id]['organization'] = 'Нет'
    # locksmith_job = types.InlineKeyboardButton(text='Слесарные работы', callback_data=f'cate_locksmith')
    # pickup_job = types.InlineKeyboardButton(text='Кузовные работы', callback_data=f'cate_pickupjob')
    # sticker = types.InlineKeyboardButton(text='Оклейка авто', callback_data=f'cate_sticker')
    # wash = types.InlineKeyboardButton(text='Мойка авто', callback_data=f'cate_wash')
    elif call.data[:4] == 'cate':
        if call.data[5:] == 'locksmith':
            text = 'Слесарные работы'
        elif call.data[5:] == 'pickupjob':
            text = 'Кузовные работы'
        elif call.data[5:] == 'sticker':
            text = 'Оклейка авто'
        else:
            text = 'Мойка авто'

        bot.send_message(call.message.chat.id,
                         "Введите описание работы")
        bot.register_next_step_handler(call.message,get_job_type, text)

    elif call.data[:5] == "admin":
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text=f"Мойщик:     <b> {user_dict[call.message.chat.id]['name']}</b>\n"
                                   f"Вид работ:     <b> {user_dict[call.message.chat.id]['job_type']}</b>\n"
                                   f"Пробег: {user_dict[call.message.chat.id]['probeg']}\n"
                                   f" Цена: <b> {user_dict[call.message.chat.id]['price']} р</b>\n"
                                   f" Заказ оплачен: <b> {user_dict[call.message.chat.id]['paid']}</b>\n"
                                   f"  Номер авто:  <b> {user_dict[call.message.chat.id]['number_auto']}</b>\n"
                                   f" Организация(при безнал оплате):  <b> {user_dict[call.message.chat.id]['organization']}</b>\n "
                                   f"Заказ верен?",
                              reply_markup=inline_keyboard_order_check, parse_mode='HTML')
        user_dict[call.message.chat.id]['to_admin'] = call.data[6:]

    elif call.data[:5] == "order":

        if call.data[12:] == 'yes':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Заказ отправлен",
                                  reply_markup=None)
            bot.send_photo(chat_id=user_dict[call.message.chat.id]['to_admin'],
                           photo=open(user_dict[call.message.chat.id]['send_photo'], 'rb'),
                           caption=f"Мойщик:     <b> {user_dict[call.message.chat.id]['name']}</b>\n"
                                   f"Вид работ:     <b> {user_dict[call.message.chat.id]['job_type']}</b>\n"
                                   f"Пробег: {user_dict[call.message.chat.id]['probeg']}\n"
                                   f" Цена: <b> {user_dict[call.message.chat.id]['price']} р</b>\n"
                                   f" Заказ оплачен: <b> {user_dict[call.message.chat.id]['paid']}</b>\n"
                                   f"  Номер авто:  <b> {user_dict[call.message.chat.id]['number_auto']}</b>\n"
                                   f" Организация(при безнал оплате):  <b> {user_dict[call.message.chat.id]['organization']}</b>\n ",
                           parse_mode='HTML')
            #/\ костыль, который попросил заказчик,\/ нормальный код
            # bot.send_photo(chat_id=572659470,
            #                photo=open(user_dict[call.message.chat.id]['send_photo'], 'rb'),
            #                caption=f"Вид работ:     <b> {user_dict[call.message.chat.id]['job_type']}</b>\n"
            #                        f" Цена: <b> {user_dict[call.message.chat.id]['price']} р</b>\n"
            #                        f" Заказ оплачен: <b> {user_dict[call.message.chat.id]['paid']}</b>\n"
            #                        f"  Номер авто:  <b> {user_dict[call.message.chat.id]['number_auto']}</b>\n"
            #                        f" Организация(при безнал оплате):  <b> {user_dict[call.message.chat.id]['organization']}</b>\n ",
            #                parse_mode='HTML')

            excel_maker(call)
        elif call.data[12:] == 'no':
            bot.send_message(chat_id=call.message.chat.id, text="Заказ удален")
    elif call.data[:3] == "reg":
        if check_id_in_file(call.message.chat.id, path_of_registration_file_admins):
            path = path_of_registration_file_admins
        else:
            path = path_of_registration_file

        if call.data[4:] == "yes":
            bot.send_message(call.message.chat.id,
                             "Спасибо, вы зарегестрированы\n Если захотите изменить данные или удалить их, то снова напишите /reg\n Для начала работы нажмите кнопку Создать заказ",
                             reply_markup=keyboard_main())
        else:
            deleter_of_data(call.message, path)
    elif call.data[:6] == "change":  # change_reg_yes
        if check_id_in_file(call.message.chat.id, path_of_registration_file_admins):
            path = path_of_registration_file_admins
        else:
            path = path_of_registration_file

        if call.data[11:] == "yes":  # change data
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Какие данные вы хотите изменить?",
                                  reply_markup=inline_keyboard_change_name)
        else:  # delete data
            deleter_of_data(call.message, path)
    elif call.data[:2] == "cn":
        if check_id_in_file(call.message.chat.id, path_of_registration_file_admins):
            path = path_of_registration_file_admins
        else:
            path = path_of_registration_file

        if call.data[3] == "n":  # name
            bot.send_message(call.message.chat.id,
                             "Введите новое имя")
            bot.register_next_step_handler(call.message, change_some_thing, 1, path)
        elif call.data[3] == "s":  # second name
            bot.send_message(call.message.chat.id,
                             "Введите новою фамилию")
            bot.register_next_step_handler(call.message, change_some_thing, 2, path)
        elif call.data[3] == "t":  # thierd name
            bot.send_message(call.message.chat.id,
                             "Введите новое отчество")
            bot.register_next_step_handler(call.message, change_some_thing, 3, path)
    elif call.data[:8] == 'del_org_':
        deleter_of_excel(call.message, 'organiztion/list_of_org.txt', call.data[8:])
    elif call.data[:4] == 'org_':
        if call.data[4:] == 'yes':
            bot.send_message(chat_id=call.message.chat.id,
                             text="Выберете текущего администратора",
                             reply_markup=admins_keyboard())
        elif call.data[4:] == 'no':
            bot.send_message(call.message.chat.id,
                             "Заказ удален", reply_markup=keyboard_main())
            user_dict[call.message.chat.id] = {'job_type': None, 'send_photo': None, 'paid': None, 'organization': None,
                                               'number_auto': None, 'price': None, 'to_admin': None, 'local_n': None,
                                               'name': None}
        else:  ###############################################################################################################################################################################################################
            if excel_check(f'organiztion/{call.data[4:].rstrip()}', call):
                user_dict[call.message.chat.id]['organization'] = call.data[4:]
            else:
                bot.send_message(call.message.chat.id,
                                 "Машина не найдена в списке. Всё равно продолжить создавать заказ?",
                                 reply_markup=inline_keyboard_org_continue)
                ###########################################################################################################
    elif call.data[:11] == 'del_worker_':
        with open('workers.txt', 'r') as f:
            lines_w = f.readlines()
            for line_w in lines_w:
                if call.data[11:] == line_w:
                    lines_w.remove(line_w)
                    break
        with open('workers.txt', 'w') as f:
            for line_w in lines_w:
                f.write(line_w)
        bot.send_message(call.message.chat.id,
                         "Сотрудник удален",
                         reply_markup=None)
    elif call.data[:7] == 'worker_':
        with open('workers.txt', 'r') as f:
            lines_w = f.readlines()
            for line_w in lines_w:
                if call.data[7:] == line_w:
                    user_dict[call.message.chat.id]['name'] = line_w
                    bot.send_message(call.message.chat.id, 'Укажите цену в рублях')
                    bot.register_next_step_handler(call.message, get_price)
        #########################################################################################################################
    elif call.data[:5] == 'spare':
        if call.data[12:] == 'yes':
            if os.path.exists('excel_files/spare.xlsx'):
                wb_obj = load_workbook('excel_files/spare.xlsx')
                sheet_obj = wb_obj.active
                num_of_cell = 1
                for cell in sheet_obj['A']:
                    if cell.value != None:
                        num_of_cell += 1
                sheet_obj[f'A{num_of_cell}'] = user_dict[call.message.chat.id]['name_spare']
                sheet_obj[f'B{num_of_cell}'] = user_dict[call.message.chat.id]['count_spare']
                sheet_obj[f'C{num_of_cell}'] = 0
                sheet_obj[f'D{num_of_cell}'] = user_dict[call.message.chat.id]['price_spare']
                sheet_obj[f'F{num_of_cell}'] = user_dict[call.message.chat.id]['count_spare']
                sheet_obj[f'G{num_of_cell}'] = datetime.datetime.today().strftime("%d.%m.%Y")
                sheet_obj[f'H{num_of_cell}'] = int(user_dict[call.message.chat.id]['count_spare'])*int(user_dict[call.message.chat.id]['price_spare'])
                sheet_obj[f'I{num_of_cell}'] = user_dict[call.message.chat.id]['brand']
                sheet_obj[f'J{num_of_cell}'] = user_dict[call.message.chat.id]['used']
                wb_obj.save('excel_files/spare.xlsx')
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Запись сделана",
                                  reply_markup=None)
        else:
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text="Запись прекращена",
                                  reply_markup=None)

    elif call.data == 'used_spare':
        user_dict[call.message.chat.id]['used'] = 'Б'
        bot.send_message(call.message.chat.id,
                         "Марка авто?'", reply_markup=car_brand())
    elif call.data == 'new_spare':
        user_dict[call.message.chat.id]['used'] = 'Н'
        bot.send_message(call.message.chat.id,
                         "Марка авто?'", reply_markup=car_brand())

    elif call.data[:5] == 'brand':
        if call.data[6:] == 'skoda':
            user_dict[call.message.chat.id]['brand'] = 'Ш'
        elif call.data[6:] == 'kia':
            user_dict[call.message.chat.id]['brand'] = 'К'
        else:
            user_dict[call.message.chat.id]['brand'] = 'Ш,К'
        bot.send_message(call.message.chat.id, "Введите количество запчастей", reply_markup=None)
        bot.register_next_step_handler(call.message, count_spare)

    elif call.data[:5] == 'xlsx_':
        bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
        if call.data[-4:] == 'name':
            bot.send_message(call.message.chat.id,
                             f"Введите ключевое слово для поиска", reply_markup=None)
        else:
            bot.send_message(call.message.chat.id,
                             f"Введите код запчасти для поиска для поиска", reply_markup=None)
        bot.register_next_step_handler(call.message, excel_search, call.data[5:-5], call.data[-5:])
    elif call.data[:9] == 'skip_row_':
        skip_list.append(call.data[9:])
        bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
        code_giver(call.message, True)
    elif call.data == "cancel":
        bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
        bot.send_message(call.message.chat.id,
                         "Для начала работы нажмите 'Создать заказ'", reply_markup=keyboard_main())


# delete file
def deleter_of_excel(message, path, name):
    with open(path, 'r') as new_file:
        list_to_delete = new_file.readlines()
        for line in list_to_delete:
            if name in line:
                try:
                    os.remove(f'organiztion/{name.rstrip()}.xlsx')
                except:
                    pass
                list_to_delete.remove(line)
    with open(path, 'w') as new_file:
        for i in list_to_delete:
            new_file.write(i + '\n')
    bot.send_message(message.chat.id,
                     "Файл удален",
                     reply_markup=None)


# delete data of user
def deleter_of_data(message, path):
    with open(path, 'r') as new_file:
        list_to_delete = new_file.readlines()
        for line in list_to_delete:
            if str(message.chat.id) in line:
                list_to_delete.remove(line)
    with open(path, 'w') as new_file:
        for i in list_to_delete:
            new_file.write(i + '\n')
    bot.send_message(message.chat.id,
                     "Данные удаленны, для повторной регистрации введите /reg",
                     reply_markup=None)


# check to id in file
def check_id_in_file(id, path):
    with open(path, 'r') as f:
        lines = f.readlines()
        for line in lines:
            if str(id) in line:
                return True
        return False


# kb for create order
def keyboard_main():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    create_order = types.KeyboardButton('Создать заказ')
    get_btn = types.KeyboardButton('Приход')
    give_btn = types.KeyboardButton('Выдача')

    markup.add(create_order)
    markup.add(get_btn, give_btn)
    return markup


# listening server
@bot.message_handler(content_types=['text', 'photo', 'document'])
# main cmds
def start(message):
    if not check_id_in_file(message.chat.id, path_of_registration_file) and not check_id_in_file(message.chat.id,
                                                                                                 path_of_registration_file_admins):
        if message.text == '/start':
            bot.send_message(message.chat.id,
                             """ 
                             Добро пожаловать, я ваш бот помощник, для начала нужно зарегестрироваться,
                             для этого введите /reg если вы сотрудник и /admin, если вы администратор.
                             """, )
        elif message.text == '/reg':
            registration_washer(message, path_of_registration_file)
        elif message.text == '/admin':
            bot.send_message(message.chat.id, 'Введите пароль администратора')
            bot.register_next_step_handler(message, admin_password_check)
        else:
            bot.send_message(message.chat.id, 'Вы не зарегестрированны')
    else:
        if message.text == 'Создать заказ':
            user_dict[message.chat.id] = {'job_type': None, 'send_photo': None, 'paid': None, 'organization': None,
                                          'number_auto': None, 'price': None, 'to_admin': None, 'local_n': None,
                                          'name': None, 'probeg': None}
            start_massage(message)
        elif message.text == '/reg':
            registration_washer(message, path_of_registration_file)
        elif message.text == '/admin':
            bot.send_message(message.chat.id, 'Введите пароль администратора')
            bot.register_next_step_handler(message, admin_password_check)
        elif message.text == 'Регистрация':
            bot.send_message(message.chat.id, 'Введите пароль администратора')
            bot.register_next_step_handler(message, path_of_registration_file_admins)
        elif message.text == 'Добавить организацию':
            bot.send_message(message.chat.id, "Пришлите excel файл организации", reply_markup=None)
            bot.register_next_step_handler(message, add_file)
        elif message.text == 'Удалить организацию':
            bot.send_message(message.chat.id, "Какую организацию вы хотите удалить?",
                             reply_markup=org_keyboard(False))
            ##################################################################################################################
        elif message.text == 'Добавить работника':
            bot.send_message(message.chat.id, "Укажите Фамилию сотрудника", reply_markup=None)
            bot.register_next_step_handler(message, add_worker)

        elif message.text == 'Удалить работника':
            bot.send_message(message.chat.id, "Какого сотрудника вы хотите удалить?",
                             reply_markup=worker_keyboard(False))
        elif message.text == 'Обновить склад':
            refresh_storage(message)
        elif message.text == 'Удалить запчасть':
            user_dict[message.chat.id] = {'name_spare': None, 'price_spare': None, 'count_spare': None, 'code': None}
            bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
            bot.send_message(message.chat.id,
                             f"Для поиска запчасти можете ввести ключевое слово или код",
                             reply_markup=spare_keyboard_get('del_spare_'))
        elif message.text == 'Вернуться':
            bot.send_message(message.chat.id, "Возращаюсь", reply_markup=keyboard_main())
        elif message.text == 'Приход':
            user_dict[message.chat.id] = {'name_spare': None, 'price_spare': None, 'count_spare': None}
            bot.send_message(message.chat.id,
                             "Введите название запчасти или воспользуйтесь поиском по ключевому слову или коду",
                             reply_markup=spare_keyboard_get('get_'))

            bot.register_next_step_handler(message, name_spare)
        elif message.text == 'Выдача':
            user_dict[message.chat.id] = {'name_spare': None, 'price_spare': None, 'count_spare': None, 'code': None,
                                          'probeg_sp': None, 'used':None, 'car_brand':None}
            bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
            bot.send_message(message.chat.id,
                             f"Для поиска запчасти можете ввести ключевое слово или код",
                             reply_markup=spare_keyboard_get('give_'))
        elif message.text == 'Ввести коды':
            code_giver(message, True)

        else:
            bot.send_message(message.chat.id, "Не понял вас", reply_markup=keyboard_main())


# check admins password
def admin_password_check(message):
    if message.text == password_of_admins:
        bot.send_message(message.chat.id, 'Вберите действие', reply_markup=kb_cmd_admin())
    else:
        bot.send_message(message.chat.id, 'Пароль не правильный')


def registration_washer(message, path):
    if_not_reg = True
    new_file = open(path, 'r')
    lines = new_file.readlines()
    for line in lines:
        if str(message.from_user.id) in line:
            bot.send_message(message.chat.id,
                             "Вы уже зарегестрированны, хотите изменить данные или удалить и совсем?",
                             reply_markup=inline_keyboard_reg_change)
            if_not_reg = False
            new_file.close()
    new_file.close()
    if if_not_reg:
        bot.send_message(message.chat.id, "Введите ваше имя", reply_markup=None)
        bot.register_next_step_handler(message, get_name_washer, path)


# get washer fio
def get_name_washer(message, path):
    with open(path, 'a') as new_file:
        new_file.write(str(message.from_user.id) + ' ' + message.text + ' ')
    bot.send_message(message.chat.id, "Введите вашу фамилию", reply_markup=None)
    bot.register_next_step_handler(message, get_second_name_washer, path)


def get_second_name_washer(message, path):
    with open(path, 'a') as new_file:
        new_file.write(message.text + ' ')
    bot.send_message(message.chat.id, "Введите ваше отчество", reply_markup=None)
    bot.register_next_step_handler(message, get_thierd_name_washer, path)


def get_thierd_name_washer(message, path):
    with open(path, 'a') as new_file:
        new_file.write(message.text + '\n')
    with open(path, 'r') as new_file:
        line = None
        while line != '':
            line = new_file.readline()
            if str(message.from_user.id) in line:
                list_of_line = list(line.split())
                bot.send_message(message.chat.id,
                                 f'Данные верны?\n{list_of_line[1]} {list_of_line[2]} {list_of_line[3]}',
                                 reply_markup=inline_keyboard_reg)  # Регистрация


def change_some_thing(message, index_in_save_list, path):
    with open(path, 'r') as f:
        save_list = f.readlines()
    for line in save_list:
        if str(message.from_user.id) in line:
            list_line = list(line.split())
            list_line[index_in_save_list] = message.text
            save_list.remove(line)
    save_list.append(f'{list_line[0]} {list_line[1]} {list_line[2]} {list_line[3]}\n')
    with open(path, 'w') as f:
        for line in save_list:
            f.write(line)

    bot.send_message(message.chat.id,
                     f'Данные изменены.\nЖелаете изменить что-то еще?',
                     reply_markup=inline_keyboard_change_name)


def add_file(message):
    if message.document != None:
        if message.document.file_name.endswith('xlsx'):
            file_info_excel = bot.get_file(message.document.file_id)
            downloaded_file_excel = bot.download_file(file_info_excel.file_path)
            with open(f'organiztion/{message.document.file_name}', 'wb') as f:
                f.write(downloaded_file_excel)
            with open(f'organiztion/list_of_org.txt', 'a') as f:
                name = f'{message.document.file_name[:-5]}\n'
                f.write(name)
            bot.send_message(message.chat.id, 'файл принят')
        else:
            bot.send_message(message.chat.id, 'Не правильный тип файла')
    else:
        bot.send_message(message.chat.id, 'нужно прислать файл')


##################################################################################################################################
def add_worker(message):
    with open('workers.txt', 'a') as f:
        f.write(message.text + '\n')
    bot.send_message(message.chat.id, 'Сотрудник добавлен')


##################################################################################################################################

def start_massage(message):
    bot.send_message(message.chat.id, "Укажите вид работ", reply_markup=get_job_category())


# get oreder data
#мне лень менять все переменные, job_type это описание работы
def get_job_type(message, category_job):
    if message.text != 'Создать заказ':
        user_dict[message.chat.id]['job_type'] = category_job
        user_dict[message.chat.id]['job_type'] =user_dict[message.chat.id]['job_type'] +': ' + message.text
        bot.send_message(message.chat.id, 'Укажите пробег')
        bot.register_next_step_handler(message, probeg)
        # bot.send_message(message.chat.id, 'Выберите себя из списка', reply_markup=worker_keyboard(True))
        # bot.send_message(message.chat.id, 'Укажите цену в рублях')
        # bot.register_next_step_handler(message, get_price)
    else:
        bot.send_message(message.chat.id, 'Заказ уже создается, укажите вид работ')
        bot.register_next_step_handler(message, get_job_type)


def probeg(message):
    user_dict[message.chat.id]['probeg'] = message.text
    bot.send_message(message.chat.id, 'Выберите себя из списка', reply_markup=worker_keyboard(True))


def get_price(message):
    if message.text != 'Создать заказ' and message.text != None:
        try:
            user_dict[message.chat.id]['price'] = int(message.text)
            bot.send_message(message.chat.id, 'Сфотографируйте автомобиль')
            bot.register_next_step_handler(message, get_photo)
        except ValueError:
            bot.send_message(message.chat.id, 'Сообщение должно содержать только цифры.\n Введите цену еще раз')
            bot.register_next_step_handler(message, get_price)
    else:
        bot.send_message(message.chat.id, 'Заказ уже создается, укажите цену')
        bot.register_next_step_handler(message, get_price)


def get_photo(message):
    with open(path_to_n_cont, 'r') as f:
        n = f.read()
        n = int(n.strip())
    if message.photo != None:
        file_info = bot.get_file(message.photo[-1].file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        src = f'photo/{date.today()}___{int(n) - 1}.jpg'
        user_dict[message.chat.id]['local_n'] = int(n) - 1
        user_dict[message.chat.id]['send_photo'] = src
        with open(src, 'wb') as new_file:
            new_file.write(downloaded_file)
        bot.send_message(message.chat.id, "Введите номер авто", reply_markup=None)
        bot.register_next_step_handler(message, get_number_auto)

    else:
        bot.send_message(message.chat.id, 'Пришлите фотографию')
        bot.register_next_step_handler(message, get_photo)


def get_number_auto(message):
    if message.text != 'Создать заказ':
        user_dict[message.chat.id]['number_auto'] = message.text
        bot.send_message(message.chat.id, 'Заказ оплачен?', reply_markup=inline_keyboard_paid)
    else:
        bot.send_message(message.chat.id, 'Заказ уже создается, укажите номер авто')
        bot.register_next_step_handler(message, get_number_auto)


# Блок обработки кнопки Выдача
def name_spare(message):
    user_dict[message.chat.id]['name_spare'] = message.text
    bot.send_message(message.chat.id, "Укажите цену запчасти в рублях", reply_markup=None)
    bot.register_next_step_handler(message, price_spare)


def price_spare(message):
    try:
        user_dict[message.chat.id]['price_spare'] = int(message.text)
    except:
        bot.send_message(message.chat.id, "Используйте только цифры, пожалуйста", reply_markup=None)
        bot.register_next_step_handler(message, price_spare)
        return
    bot.send_message(message.chat.id, "Запчасть б/у?", reply_markup=bu())


def count_spare(message):
    try:
        user_dict[message.chat.id]['count_spare'] = int(message.text)
    except:
        bot.send_message(message.chat.id, "Используйте только цифры, пожалуйста", reply_markup=None)
        bot.register_next_step_handler(message, count_spare)
        return
    bot.send_message(message.chat.id, f"{user_dict[message.chat.id]['name_spare']}\n"
                                      f"{user_dict[message.chat.id]['used']}   "
                                      f"{user_dict[message.chat.id]['brand']} \n "
                                      f"{user_dict[message.chat.id]['count_spare']} шт   "
                                      f"{user_dict[message.chat.id]['price_spare']}р.", reply_markup=spare_confirm)

def get_plus(message, row, plus):

    try:
        new_price = int(message.text)
    except:
        bot.send_message(message.chat.id, "Используйте только цифры, пожалуйста", reply_markup=None)
        bot.register_next_step_handler(message, get_plus, row, plus)
    wb = load_workbook('excel_files/spare.xlsx')
    ws = wb.active
    ws[f'D{row[0].row}'] = new_price
    ws[f'B{row[0].row}'] = int(ws[f'B{row[0].row}'].value) + plus

    ws[f'F{row[0].row}'] = int(ws[f'F{row[0].row}'].value) + plus
    ws[f'H{row[0].row}'] = int(ws[f'H{row[0].row}'].value) + plus*int(ws[f'D{row[0].row}'].value)
    ws[f'G{row[0].row}'] = datetime.datetime.today().strftime("%d.%m.%Y")
    bot.send_message(message.chat.id,
                     f"{row[0].value} {ws[f'F{row[0].row}'].value}/{ws[f'B{row[0].row}'].value} шт\n"
                     f"По цене {ws[f'D{row[0].row}'].value} р",
                     reply_markup=keyboard_main())
    wb.save('excel_files/spare.xlsx')


def give_spare(message, row):
    try:
        negative = int(message.text)
    except:
        bot.send_message(message.chat.id, "Используйте только цифры, пожалуйста", reply_markup=None)
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        bot.register_next_step_handler(message, give_spare, row)
        return
    wb = load_workbook('excel_files/spare.xlsx')
    ws = wb.active
    if int(ws[f'F{row[0].row}'].value) - negative < 0:
        bot.send_message(message.chat.id, "На складе нет такого количества запчастей", reply_markup=None)
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        bot.register_next_step_handler(message, give_spare, row)
        return
    else:
        ws[f'C{row[0].row}'] = int(ws[f'C{row[0].row}'].value) + negative
        ws[f'F{row[0].row}'] = int(ws[f'F{row[0].row}'].value) - negative
        ws[f'H{row[0].row}'] = int(ws[f'H{row[0].row}'].value) - negative*int(ws[f'D{row[0].row}'].value)
        bot.send_message(message.chat.id, "Введите пробег", reply_markup=None)
        bot.register_next_step_handler(message, probeg_sp, row, negative)
    wb.save('excel_files/spare.xlsx')


def probeg_sp(message, row, negative):
    user_dict[message.chat.id]['probeg_sp'] = message.text
    bot.send_message(message.chat.id, "Введите номер машины, водителю которой выданы запчасти", reply_markup=None)
    bot.register_next_step_handler(message, num_spare, row, negative)


def num_spare(message, row, negative):
    wb = load_workbook('excel_files/spare.xlsx')
    ws = wb.active
    num_spare_ = message.text
    rows = ws['M']
    row_count = 1
    for cell in rows:
        if cell.value != None:
            row_count += 1
    ws[f'M{row_count}'] = num_spare_
    ws[f'N{row_count}'] = row[0].value
    ws[f'O{row_count}'] = negative


    try:
        ws[f'P{row_count}'] = negative * int(ws[f'D{row[0].row}'].value) + int(ws[f'D{row[0].row}'].value) * 0.2*negative
    except:
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        bot.send_message(message.chat.id, 'Ошибка, попробуйте снова')
        return
    ws[f'Q{row_count}'] = user_dict[message.chat.id]['probeg_sp']
    ws[f'R{row_count}'] = datetime.datetime.today().strftime("%d.%m.%Y")

    bot.send_message(message.chat.id,
                     f"{num_spare_}\n"
                     f"{row[0].value} выдано {negative} шт\n"
                     f"Общая стоимость выданных деталей: {ws[f'P{row_count}'].value} р.\n"
                     f"На складе осталось {ws[f'F{row[0].row}'].value}/{ws[f'B{row[0].row}'].value} шт",
                     reply_markup=keyboard_main())
    wb.save('excel_files/spare.xlsx')


def encode_me(name):
    # open xlsx in binary mode
    with open(name, "rb") as attachment:
        # head message application/octet-stream
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        # encode message to ASCII for send
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {name}",
    )
    return part


def excel_search(message, way, search_setting):
    search_spare = message.text
    wb = load_workbook('excel_files/spare.xlsx')
    ws = wb.active
    row_nums = {}
    i = 1
    spare_list = ''
    if search_setting == '_name':
        for row in ws.iter_rows(2, ws.max_row):
            if row[0].value == None:
                continue
            if search_spare.lower() in row[0].value.lower():
                row_nums[i] = row[0].row
                spare_list = spare_list + f"{i}. {row[0].value.capitalize()}\n"
                i = i + 1
    else:
        for row in ws.iter_rows(2, ws.max_row):
            if row[4].value == None:
                continue
            if search_spare.lower() in str(row[4].value).lower():
                row_nums[0] = row[4].row
                bot.send_message(message.chat.id, 'Найденная запчасть:\n'
                                                  f'{row[0].value} {row[5].value}/{row[1].value} шт', reply_markup=None)
                spare_list = '1'

                return search_choice(message, way, row_nums, True)
    if spare_list == '':
        bot.send_message(message.chat.id, 'Ничего подобного не нашлось', reply_markup=keyboard_main())
        return
    bot.send_message(message.chat.id, f'Введите цифру нужной запчасти\n\n{spare_list}')
    bot.register_next_step_handler(message, search_choice, way, row_nums)

def get_new_price(message, row, now_price):
    try:
        plus = int(message.text)
    except:
        bot.send_message(message.chat.id, "Используйте только цифры, пожалуйста", reply_markup=None)
        bot.register_next_step_handler(message, get_new_price, row)
        return
    bot.send_message(message.chat.id, f"Сейчас цена деталей {now_price}\n"
                                      "Какую цену деталей сделать?", reply_markup=None)
    bot.register_next_step_handler(message, get_plus, row, plus)

def search_choice(message, way, row_nums,code=False):

    try:
        if code:
            current_row_spare = row_nums[0]
        else:
            current_row_spare = int(row_nums[int(message.text)])
    except:
        bot.send_message(message.chat.id, f'Ошибка, повторите попытку')
        bot.register_next_step_handler(message, way, row_nums)
        return
    # del_spare_ give_ get_
    if way == 'get_':
        wb_get = load_workbook('excel_files/spare.xlsx')
        ws_get = wb_get.active
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        bot.send_message(message.chat.id,
                         f"Сейчас на складе {ws_get[f'F{current_row_spare}'].value}/{ws_get[f'B{current_row_spare}'].value} шт\n "
                         "Сколько запчастей добавить?", reply_markup=None)
        # bot.register_next_step_handler(message, get_plus, ws_get[f'{current_row_spare}'])
        bot.register_next_step_handler(message, get_new_price, ws_get[f'{current_row_spare}'], ws_get[f'D{current_row_spare}'].value)
        wb_get.save('excel_files/spare.xlsx')

    elif way == 'give_':
        wb_give = load_workbook('excel_files/spare.xlsx')
        ws_give = wb_give.active
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        bot.send_message(message.chat.id,
                         f"Сейчас на складе {ws_give[f'F{current_row_spare}'].value}/{ws_give[f'B{current_row_spare}'].value} шт\n "
                         "Сколько запчастей выдать?", reply_markup=None)
        bot.register_next_step_handler(message, give_spare, ws_give[f'{current_row_spare}'])
        wb_give.save('excel_files/spare.xlsx')

    elif way == 'del_spare_':
        wb_del = load_workbook('excel_files/spare.xlsx')
        ws_del = wb_del.active
        bot.send_message(message.chat.id,
                         f"{ws_del[f'A{current_row_spare}'].value} {ws_del[f'F{current_row_spare}'].value}/{ws_del[f'B{current_row_spare}'].value} шт\n"
                         f"Удалено")

        ws_del.move_range(f"A{current_row_spare + 1}:G9999", rows=-1, cols=0)
        wb_del.save('excel_files/spare.xlsx')


def code_giver(message, first_call=False, row_num=None):
    global skip_list
    wb = load_workbook('excel_files/spare.xlsx')
    ws = wb.active
    if first_call != True:
        ws[f'E{row_num}'] = message.text
    for row in ws.iter_rows(2, ws.max_row):
        if str(row[0].row) in skip_list:
            continue
        if row[4].value == None and row[0].value != None:
            bot.send_message(message.chat.id, f'Впишите код для запчасти:\n'
                                              f'{row[0].value}', reply_markup=code_giver_kb(row[0].row))
            bot.register_next_step_handler(message,code_giver, row_num = row[0].row)
            wb.save('excel_files/spare.xlsx')
            return
    wb.save('excel_files/spare.xlsx')
    bot.send_message(message.chat.id, f'Нет запчастей без кода')
    skip_list = []




def refresh_storage(tlg_message):
    date_today = str(date.today())
    os.rename('excel_files/spare.xlsx', f'storage_{date_today}.xlsx')

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(bot_email, bot_password)

    # setting letter
    message = MIMEMultipart()
    message["From"] = bot_email
    message["To"] = s_k_email
    message["Subject"] = f"Отчет о складе за {date_today}"

    message.attach(MIMEText(f'Отчет за {date_today}', "plain"))
    message.attach(encode_me(f'storage_{date_today}.xlsx'))
    text = message.as_string()

    server.sendmail(bot_email, s_k_email, text)
    del text
    del message

    server.quit()

    wb_old = load_workbook(f'storage_{date_today}.xlsx')
    ws_old = wb_old.active

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Наименование'
    ws['B1'] = 'Пришло на склад'
    ws['C1'] = 'Отдано'
    ws['D1'] = 'Цена'
    ws['E1'] = 'Код'
    ws['F1'] = 'Остаток'
    ws['GH'] = 'б/у'
    ws['GI'] = 'Марка'
    ws['GJ'] = 'Итого'



    ws['M1'] = 'Номер'
    ws['N1'] = 'Запчасть'
    ws['O1'] = 'Кол-во'
    ws['P1'] = 'Общая сумма'
    ws['Q1'] = 'Пробег'
    ws['R1'] = 'Дата'
    for row in ws_old.iter_rows(2):
        if row[0].value == None:
            continue
        ws[f'A{row[0].row}'] = row[0].value
        ws[f'B{row[1].row}'] = int(row[1].value) - int(row[2].value)
        ws[f'C{row[2].row}'] = 0
        ws[f'D{row[3].row}'] = row[3].value
        ws[f'E{row[3].row}'] = row[4].value
        ws[f'F{row[3].row}'] = row[5].value
    wb.save('excel_files/spare.xlsx')
    wb_old.save(f'storage_{date_today}.xlsx')
    os.remove(f'storage_{date_today}.xlsx')
    bot.send_message(tlg_message.chat.id, 'Склад обновлен')


if __name__ == '__main__':
    while True:
        try:
            bot.polling(none_stop=True)

        except Exception as e:
            traceback.print_exc()
            time.sleep(5)
