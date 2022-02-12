import traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import pprint
import time

wb_sp = load_workbook('excel_files/spare.xlsx')
ws_sp = wb_sp.active
same_spare_dict = {}
for row in ws_sp.iter_rows(2, ws_sp.max_row):
    try:
        if row[4].value == None:
            ws_sp[f'E{row[0].row}'] = f'{row[0].row}_{row[0].value[:2]}'
        if same_spare_dict.setdefault(row[4].value, ) == None:
            same_spare_dict[row[4].value] = {}
        same_spare_dict[row[4].value]['name'] = row[0].value
        same_spare_dict[row[4].value]['arrive'] = same_spare_dict[row[4].value].setdefault('arrive', 0) + int(
            row[1].value)
        same_spare_dict[row[4].value]['give'] = same_spare_dict[row[4].value].setdefault('give', 0) + int(row[2].value)
        if same_spare_dict[row[4].value].setdefault('price', 0) < float(row[3].value):
            same_spare_dict[row[4].value]['price'] = float(row[3].value)
        same_spare_dict[row[4].value]['code'] = row[4].value
        same_spare_dict[row[4].value]['sklad'] = same_spare_dict[row[4].value].setdefault('sklad', 0) + int(
            row[5].value)
        if same_spare_dict[row[4].value].setdefault('date', datetime(year=2000, month=1, day=1, hour=0, minute=0,
                                                                     second=0)).timestamp() < row[6].value.timestamp():
            same_spare_dict[row[4].value]['date'] = row[6].value
    except Exception as e:
        traceback.print_exc()
        print(row[0].row)
wb_sp.save('excel_files/spare.xlsx')

wb_save = load_workbook('excel_files/spare_save.xlsx')
ws_save = wb_save.active
row = 2
for key in same_spare_dict:
    ws_save[f'A{row}'] = same_spare_dict[key]['name']
    ws_save[f'B{row}'] = same_spare_dict[key]['arrive']
    ws_save[f'C{row}'] = same_spare_dict[key]['give']
    ws_save[f'D{row}'] = same_spare_dict[key]['price']
    ws_save[f'E{row}'] = same_spare_dict[key]['code']
    ws_save[f'F{row}'] = same_spare_dict[key]['sklad']
    ws_save[f'G{row}'] = same_spare_dict[key]['date']
    row = row + 1
wb_save.save('excel_files/spare_save.xlsx')

wb_new = load_workbook('excel_files/spare_new.xlsx')
ws_new = wb_new.active
row_ = 2
for keyy in same_spare_dict:
    ws_new[f'A{row_}'] = same_spare_dict[keyy]['name']
    ws_new[f'B{row_}'] = same_spare_dict[keyy]['arrive']
    ws_new[f'C{row_}'] = same_spare_dict[keyy]['give']
    ws_new[f'D{row_}'] = same_spare_dict[keyy]['price']
    try:
        if "_" in same_spare_dict[keyy]['code']:
            ws_new[f'E{row_}'] = None
        else:
            ws_new[f'E{row_}'] = same_spare_dict[keyy]['code']
    except:
        ws_new[f'E{row_}'] = same_spare_dict[keyy]['code']
    ws_new[f'F{row_}'] = same_spare_dict[keyy]['sklad']
    ws_new[f'G{row_}'] = same_spare_dict[keyy]['date'].date()
    row_ = row_ + 1
wb_new.save('excel_files/spare_new.xlsx')
