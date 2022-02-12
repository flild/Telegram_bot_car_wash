import schedule
import time
import telebot
import zipfile
import os
import shutil


from datetime import date, timedelta
from config import token_b, s_k_id, path_dir
from openpyxl import Workbook
from openpyxl import load_workbook
try:
    os.chdir(path_dir)
except:
    pass
path_to_n_cont = 'n_count.txt'

def zeroing_data():
    date_yesterday = str(date.today() - timedelta(days=1))
    os.remove(f'excel_files/{date_yesterday}.xlsx')
    os.remove(f'excel_files/report.xlsx')
    os.remove(f'{date_yesterday}_photo.zip')
    os.remove(f'excel_files/spare_{date_yesterday}.xlsx')

    shutil.rmtree('photo')
    os.mkdir('photo')
    with open(path_to_n_cont, 'w') as f:
        f.write(str(2))
    pass


def job():
    try:
        bot = telebot.TeleBot(token_b)
        date_yesterday = str(date.today() - timedelta(days=1))


        if os.path.exists('excel_files/report.xlsx'):
            with open(path_to_n_cont, 'r') as f:
                n = int(f.read().strip())
            wb_obj = load_workbook('excel_files/report.xlsx')
            sheet_obj = wb_obj.active
            sheet_obj[f'G{n}'] = f'= sum(G2:G{n-1})'
            sheet_obj[f'F{n}'] = 'Сумма заказов:'
            wb_obj.save(f'excel_files/{date_yesterday}.xlsx')
            bot.send_message(s_k_id, 'Доброе утро')
            bot.send_document(s_k_id, data=open(f'excel_files/{date_yesterday}.xlsx', 'rb'),
                              caption=f'Отчет за {date_yesterday}')

            fantasy_zip = zipfile.ZipFile(f'{date_yesterday}_photo.zip', 'w')

            for folder, subfolders, files in os.walk('photo'):

                for file in files:
                    if file.endswith('.jpg'):
                        fantasy_zip.write(os.path.join(folder, file),
                                          os.path.relpath(os.path.join(folder, file), 'photo'),
                                          compress_type=zipfile.ZIP_DEFLATED)

            fantasy_zip.close()
            bot.send_document(s_k_id, data=open(f'{date_yesterday}_photo.zip', 'rb'),
                              caption=f'фото машин за  {date.today() - timedelta(days=1)}')
            del fantasy_zip
            del files
            del folder
            del wb_obj

        if os.path.exists('excel_files/spare.xlsx'):

            os.rename('excel_files/spare.xlsx', f'excel_files/spare_{date_yesterday}.xlsx')
            wb_sp = load_workbook(f'excel_files/spare_{date_yesterday}.xlsx')
            sheet_sp = wb_sp.active
            wb_sp.save(f'excel_files/spare.xlsx')

            bot.send_document(s_k_id, data=open(f'excel_files/spare_{date_yesterday}.xlsx', 'rb'),
                              caption=f'Отчет запчастей за {date.today() - timedelta(days=1)}')
            del wb_sp
            zeroing_data()
    except Exception as es:
        with open('error.txt', 'a') as f:
            f.write(str(es))


schedule.every().day.at("07:00").do(job)

while True:
    schedule.run_pending()
    txt = input()
    if txt == 'send':
        job()
    time.sleep(60)
